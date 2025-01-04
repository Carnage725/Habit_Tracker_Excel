from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import calendar
import os

# Prompt the user for number of habits and their names
num_habits = int(input("Enter the number of habits: "))
habit_names = [input(f"Enter habit #{i + 1}: ") for i in range(num_habits)]

# Habit Tracker Template Data
habit_tracker_template = {
    "Habit": habit_names,
    "Completed? (Y/N)": ["" for _ in range(num_habits)],
    "Notes": ["" for _ in range(num_habits)],
}

# Month names
months = [calendar.month_name[i] for i in range(1, 13)]

# Create an output folder for the Excel files
output_folder = "templates_output"
os.makedirs(output_folder, exist_ok=True)

# Create an Excel file for each month
for month in months:
    # Determine the number of days in the month (assuming non-leap year for simplicity)
    days_in_month = calendar.monthrange(2024, months.index(month) + 1)[1]
    
    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Create sheets for each day in the month
    for day in range(1, days_in_month + 1):
        sheet_name = f"Day {day}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Add headers
        headers = ["Habit", "Completed? (Y/N)", "Notes"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add habit tracker data
        for row_idx, habit in enumerate(habit_tracker_template["Habit"], start=2):
            ws.cell(row=row_idx, column=1, value=habit)
            ws.cell(row=row_idx, column=2, value="")
            ws.cell(row=row_idx, column=3, value="")
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
    
    # Save the workbook
    output_file = os.path.join(output_folder, f"{month}.xlsx")
    wb.save(output_file)
    print(f"Habit tracker template saved as {output_file}")
